from openpyxl import load_workbook
from apps.clientes.models import Clientes
from apps.productos.models import Producto
from decimal import Decimal, InvalidOperation
from django.db import transaction
from cities_light.models import City
from apps.proveedores.models import Proveedor
from apps.proyectos.models import Proyectos
from apps.ordenes_compra.models import OrdenCompra, ProductoSolicitado
from datetime import datetime
from dateutil import parser
import locale


def procesar_cargue_productos(archivo):
    """
    Procesa el archivo Excel de productos y crea registros nuevos.
    Si encuentra un producto existente o hay errores, cancela todo el cargue.
    Autor: Jeison Acevedo
    """

    resultados = {
        'exitosos': 0,
        'fallidos': 0,
        'errores': []
    }

    try:
        wb = load_workbook(archivo)
        ws = wb.active

        productos_a_crear = []

        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(fila):
                    continue

                referencia = str(fila[0]).strip() if fila[0] else None
                articulo = str(fila[1]).strip() if fila[1] else None
                precio_costo = fila[2]
                precio_venta = fila[3]
                linea = str(fila[4]).strip().upper() if fila[4] else None
                descripcion = str(fila[5]).strip() if fila[5] else None

                if not referencia:
                    resultados['errores'].append(f"Fila {fila_num}: Referencia vacía")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                if not articulo:
                    resultados['errores'].append(f"Fila {fila_num}: Artículo vacío para referencia '{referencia}'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                if not linea:
                    resultados['errores'].append(f"Fila {fila_num}: Línea vacía para referencia '{referencia}'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                if not descripcion:
                    resultados['errores'].append(f"Fila {fila_num}: Descripción vacía para referencia '{referencia}'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                if Producto.objects.filter(referencia=referencia).exists():
                    resultados['errores'].append(f"Fila {fila_num}: El producto con referencia '{referencia}' ya existe")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                try:
                    precio_costo = Decimal(str(precio_costo))
                    precio_venta = Decimal(str(precio_venta))
                except (InvalidOperation, ValueError, TypeError):
                    resultados['errores'].append(f"Fila {fila_num}: Precios inválidos para {referencia}")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                if precio_costo < 0 or precio_venta < 0:
                    resultados['errores'].append(f"Fila {fila_num}: Los precios deben ser positivos para {referencia}")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                lineas_validas = ['HOMBRE', 'DAMA', 'UNISEX']
                if linea not in lineas_validas:
                    resultados['errores'].append(f"Fila {fila_num}: Línea '{linea}' no válida para {referencia}. Valores válidos: {', '.join(lineas_validas)}")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                productos_a_crear.append(Producto(
                    referencia=referencia,
                    articulo=articulo,
                    precio_costo=precio_costo,
                    precio_venta=precio_venta,
                    linea=linea,
                    descripcion=descripcion
                ))

            except Exception as e:
                resultados['errores'].append(f"Fila {fila_num}: Error procesando - {str(e)}")
                resultados['fallidos'] += 1
                wb.close()
                return resultados

        wb.close()

        if productos_a_crear:
            try:
                with transaction.atomic():
                    Producto.objects.bulk_create(productos_a_crear)
                    resultados['exitosos'] = len(productos_a_crear)
            except Exception as e:
                resultados['errores'].append(f"Error al guardar productos en la base de datos: {str(e)}")
                resultados['fallidos'] = len(productos_a_crear)
                return resultados

    except Exception as e:
        resultados['errores'].append(f"Error al leer el archivo: {str(e)}")
        resultados['fallidos'] += 1

    return resultados

def procesar_cargue_clientes(archivo):
    """
    Procesa el archivo Excel de clientes y crea registros nuevos.
    Si encuentra un cliente existente o hay errores, cancela todo el cargue.
    Autor: Jeison Acevedo
    """

    wb = load_workbook(archivo)
    ws = wb.active
    resultados = {
        'exitosos': 0,
        'fallidos': 0,
        'errores': []
    }
    clientes_a_crear = []
    try:
        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(fila):
                    continue
                nombre_cliente = str(fila[0]).strip() if fila[0] else None
                email_cliente = str(fila[1]).strip() if fila[1] else None
                nombre_contacto = str(fila[2]).strip() if fila[2] else None
                telefono_contacto = str(fila[3]).strip() if fila[3] else None
                email_contacto = str(fila[4]).strip() if fila[4] else None
                direccion = str(fila[5]).strip() if fila[5] else None
                ciudad_name = str(fila[6]).strip() if fila[6] else None
                tiene_proyectos = str(fila[7]).strip().upper() if fila[7] else None

                if not nombre_cliente:
                    resultados['errores'].append(f"Fila {fila_num}: Nombre de cliente vacío")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados
                if Clientes.objects.filter(nombre_cliente=nombre_cliente).exists():
                    resultados['errores'].append(f"Fila {fila_num}: El cliente '{nombre_cliente}' ya existe")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados
                ciudad_obj = City.objects.filter(name=ciudad_name).first() if ciudad_name else None
                if not ciudad_obj:
                    resultados['errores'].append(f"Fila {fila_num}: Ciudad '{ciudad_name}' no encontrada")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados
                if tiene_proyectos not in ['SI', 'NO']:
                    resultados['errores'].append(f"Fila {fila_num}: Valor de proyectos inválido '{tiene_proyectos}' (debe ser SI o NO)")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados
                tiene_proyectos_bool = True if tiene_proyectos == 'SI' else False

                clientes_a_crear.append(Clientes(
                    nombre_cliente=nombre_cliente,
                    email_cliente=email_cliente,
                    nombre_contacto=nombre_contacto,
                    telefono_contacto=telefono_contacto,
                    email_contacto=email_contacto,
                    direccion_cliente=direccion,
                    ciudad_cliente=ciudad_obj,
                    tiene_proyectos=tiene_proyectos_bool
                ))
            except Exception as e:
                resultados['errores'].append(f"Fila {fila_num}: Error procesando - {str(e)}")
                resultados['fallidos'] += 1
                wb.close()
                return resultados
        wb.close()
        if clientes_a_crear:
            try:
                with transaction.atomic():
                    Clientes.objects.bulk_create(clientes_a_crear)
                    resultados['exitosos'] = len(clientes_a_crear)
            except Exception as e:
                resultados['errores'].append(f"Error al guardar clientes en la base de datos: {str(e)}")
                resultados['fallidos'] = len(clientes_a_crear)
                return resultados
    except Exception as e:
        resultados['errores'].append(f"Error al leer el archivo: {str(e)}")
        resultados['fallidos'] += 1
    return resultados

def procesar_cargue_proyectos(archivo):
    """
    Procesa el archivo Excel de proyectos y crea registros nuevos.
    Si encuentra un proyecto existente o hay errores, cancela todo el cargue.
    Autor: Jeison Acevedo
    """

    wb = load_workbook(archivo)
    ws = wb.active
    resultados = {
        'exitosos': 0,
        'fallidos': 0,
        'errores': []
    }
    proyectos_a_crear = []
    try:
        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(fila):
                    continue
                nombre_cliente = str(fila[0]).strip().lower() if fila[0] else None
                ciudad_name = str(fila[1]).strip().lower() if fila[1] else None
                nombre_proyecto = str(fila[2]).strip() if fila[2] else None
                email_proyecto = str(fila[3]).strip() if fila[3] else None
                nombre_contacto = str(fila[4]).strip() if fila[4] else None
                telefono_contacto = str(fila[5]).strip() if fila[5] else None
                email_contacto = str(fila[6]).strip() if fila[6] else None
                direccion = str(fila[7]).strip() if fila[7] else None

                campos = [nombre_cliente, ciudad_name, nombre_proyecto, email_proyecto, nombre_contacto, telefono_contacto, email_contacto, direccion]
                if not all(campos):
                    resultados['errores'].append(f"Fila {fila_num}: Algún campo está vacío")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                cliente_obj = None
                for c in Clientes.objects.filter(tiene_proyectos=True):
                    if c.nombre_cliente and c.nombre_cliente.strip().lower() == nombre_cliente:
                        cliente_obj = c
                        break
                if not cliente_obj:
                    resultados['errores'].append(f"Fila {fila_num}: Cliente '{nombre_cliente}' no existe o no tiene proyectos habilitados")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                ciudad_obj = None
                for city in City.objects.all():
                    if city.name and city.name.strip().lower() == ciudad_name:
                        ciudad_obj = city
                        break
                if not ciudad_obj:
                    resultados['errores'].append(f"Fila {fila_num}: Ciudad '{ciudad_name}' no encontrada")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                proyectos_a_crear.append(Proyectos(
                    cliente=cliente_obj,
                    ciudad_proyecto=ciudad_obj,
                    nombre_proyecto=nombre_proyecto,
                    email_proyecto=email_proyecto,
                    nombre_contacto=nombre_contacto,
                    telefono_contacto=telefono_contacto,
                    email_contacto=email_contacto,
                    direccion_proyecto=direccion
                ))
            except Exception as e:
                resultados['errores'].append(f"Fila {fila_num}: Error procesando - {str(e)}")
                resultados['fallidos'] += 1
                wb.close()
                return resultados
        wb.close()
        if proyectos_a_crear:
            try:
                with transaction.atomic():
                    Proyectos.objects.bulk_create(proyectos_a_crear)
                    resultados['exitosos'] = len(proyectos_a_crear)
            except Exception as e:
                resultados['errores'].append(f"Error al guardar proyectos en la base de datos: {str(e)}")
                resultados['fallidos'] = len(proyectos_a_crear)
                return resultados
    except Exception as e:
        resultados['errores'].append(f"Error al leer el archivo: {str(e)}")
        resultados['fallidos'] += 1
    return resultados

def procesar_cargue_proveedores(archivo):
    """
    Procesa el archivo Excel de proveedores y crea registros nuevos.
    Si encuentra un proveedor existente o hay errores, cancela todo el cargue.
    Autor: Jeison Acevedo
    """

    wb = load_workbook(archivo)
    ws = wb.active
    resultados = {
        'exitosos': 0,
        'fallidos': 0,
        'errores': []
    }
    proveedores_a_crear = []
    try:
        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(fila):
                    continue
                nombre_original = str(fila[0]).strip() if fila[0] else None
                nombre = nombre_original.lower() if nombre_original else None
                nit = str(fila[1]).strip() if fila[1] else None
                correo = str(fila[2]).strip() if fila[2] else None
                asesor_contacto = str(fila[3]).strip() if fila[3] else None
                telefono = str(fila[4]).strip() if fila[4] else None
                productos_suministra = str(fila[5]).strip() if fila[5] else None

                campos = [nombre, nit, correo, asesor_contacto, telefono, productos_suministra]
                if not all(campos):
                    resultados['errores'].append(f"Fila {fila_num}: Algún campo está vacío")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                for p in Proveedor.objects.all():
                    if p.nombre and p.nombre.strip().lower() == nombre:
                        resultados['errores'].append(f"Fila {fila_num}: El proveedor con nombre '{nombre_original}' ya existe")
                        resultados['fallidos'] += 1
                        wb.close()
                        return resultados
                if Proveedor.objects.filter(nit=nit).exists():
                    resultados['errores'].append(f"Fila {fila_num}: El proveedor con NIT '{nit}' ya existe")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                if '-' not in nit:
                    resultados['errores'].append(f"Fila {fila_num}: El NIT '{nit}' debe contener un guión (-)")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados
                partes_nit = nit.split('-')
                if len(partes_nit) != 2 or not partes_nit[0].isdigit() or not partes_nit[1].isdigit():
                    resultados['errores'].append(f"Fila {fila_num}: El NIT '{nit}' debe ser numérico en ambas partes separadas por guión")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                proveedores_a_crear.append(Proveedor(
                    nombre=nombre_original,
                    nit=nit,
                    correo=correo,
                    asesor_contacto=asesor_contacto,
                    telefono=telefono,
                    productos_suministra=productos_suministra
                ))
            except Exception as e:
                resultados['errores'].append(f"Fila {fila_num}: Error procesando - {str(e)}")
                resultados['fallidos'] += 1
                wb.close()
                return resultados
        wb.close()
        if proveedores_a_crear:
            try:
                with transaction.atomic():
                    Proveedor.objects.bulk_create(proveedores_a_crear)
                    resultados['exitosos'] = len(proveedores_a_crear)
            except Exception as e:
                resultados['errores'].append(f"Error al guardar proveedores en la base de datos: {str(e)}")
                resultados['fallidos'] = len(proveedores_a_crear)
                return resultados
    except Exception as e:
        resultados['errores'].append(f"Error al leer el archivo: {str(e)}")
        resultados['fallidos'] += 1
    return resultados


def procesar_cargue_ordenes_compra(archivo):
    """
    Procesa el archivo Excel de órdenes de compra y crea registros nuevos.
    Si encuentra una orden existente o hay errores, cancela todo el cargue.
    Autor: Jeison Acevedo
    """

    resultados = {
        'exitosos': 0,
        'fallidos': 0,
        'errores': []
    }

    try:
        wb = load_workbook(archivo)
        ws = wb.active

        ordenes_dict = {}

        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except:
            try:
                locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')
            except:
                pass

        meses_esp = {
            'ene': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'abr': 'Apr',
            'may': 'May', 'jun': 'Jun', 'jul': 'Jul', 'ago': 'Aug',
            'sep': 'Sep', 'oct': 'Oct', 'nov': 'Nov', 'dic': 'Dec'
        }

        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(fila):
                    continue

                # A: CODIGO OC
                codigo_oc = str(fila[0]).strip() if fila[0] else None
                if not codigo_oc:
                    resultados['errores'].append(f"Fila {fila_num}: Código de OC es obligatorio")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                # B: FECHA SOLICITUD
                fecha_str = str(fila[1]).strip() if fila[1] else None
                if not fecha_str:
                    resultados['errores'].append(f"Fila {fila_num}: Fecha de solicitud es obligatoria para OC '{codigo_oc}'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                try:
                    fecha_lower = fecha_str.lower()
                    for mes_esp, mes_eng in meses_esp.items():
                        fecha_lower = fecha_lower.replace(mes_esp, mes_eng)

                    try:
                        fecha_solicitud = parser.parse(fecha_lower, dayfirst=True).date()
                    except:
                        fecha_solicitud = datetime.strptime(fecha_str, '%d-%m-%Y').date()
                except Exception as e:
                    resultados['errores'].append(f"Fila {fila_num}: Formato de fecha inválido '{fecha_str}' para OC '{codigo_oc}'. Use formato como '21-dic-2021' o '21-12-2021'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                # C: CLIENTE
                nombre_cliente = str(fila[2]).strip() if fila[2] else None
                if not nombre_cliente:
                    resultados['errores'].append(f"Fila {fila_num}: Cliente es obligatorio para OC '{codigo_oc}'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                try:
                    cliente = Clientes.objects.get(nombre_cliente__iexact=nombre_cliente)
                except Clientes.DoesNotExist:
                    resultados['errores'].append(f"Fila {fila_num}: El cliente '{nombre_cliente}' no existe.")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                # D: PROYECTO
                nombre_proyecto = str(fila[3]).strip() if fila[3] else None
                proyecto = None

                if cliente.tiene_proyectos:
                    if not nombre_proyecto:
                        resultados['errores'].append(f"Fila {fila_num}: El cliente '{nombre_cliente}' requiere un proyecto asignado")
                        resultados['fallidos'] += 1
                        wb.close()
                        return resultados

                    try:
                        proyecto = Proyectos.objects.get(nombre_proyecto__iexact=nombre_proyecto, cliente=cliente)
                    except Proyectos.DoesNotExist:
                        resultados['errores'].append(f"Fila {fila_num}: El proyecto '{nombre_proyecto}' no está registrado para el cliente '{nombre_cliente}'")
                        resultados['fallidos'] += 1
                        wb.close()
                        return resultados

                # E: PRODUCTO
                referencia_producto = str(fila[4]).strip() if fila[4] else None
                if not referencia_producto:
                    resultados['errores'].append(f"Fila {fila_num}: Referencia de producto es obligatoria para OC '{codigo_oc}'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                try:
                    producto = Producto.objects.get(referencia__iexact=referencia_producto)
                except Producto.DoesNotExist:
                    resultados['errores'].append(f"Fila {fila_num}: El producto con referencia '{referencia_producto}' no existe.")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                # F: DESCRIPCIÓN EN OC
                descripcion = str(fila[5]).strip() if fila[5] else None
                if not descripcion:
                    resultados['errores'].append(f"Fila {fila_num}: Descripción es obligatoria para OC '{codigo_oc}' producto '{referencia_producto}'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                # G: CANTIDAD
                cantidad = fila[6]
                if not cantidad:
                    resultados['errores'].append(f"Fila {fila_num}: Cantidad es obligatoria para OC '{codigo_oc}' producto '{referencia_producto}'")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                try:
                    cantidad = int(cantidad)
                    if cantidad <= 0:
                        raise ValueError
                except (ValueError, TypeError):
                    resultados['errores'].append(f"Fila {fila_num}: Cantidad inválida '{cantidad}' para OC '{codigo_oc}' producto '{referencia_producto}'. Debe ser un número positivo")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                # Agrupar productos por orden de compra
                if codigo_oc not in ordenes_dict:
                    ordenes_dict[codigo_oc] = {
                        'cliente': cliente,
                        'proyecto': proyecto,
                        'fecha_solicitud': fecha_solicitud,
                        'productos': []
                    }

                ordenes_dict[codigo_oc]['productos'].append({
                    'producto': producto,
                    'cantidad': cantidad,
                    'descripcion': descripcion
                })

            except Exception as e:
                resultados['errores'].append(f"Fila {fila_num}: Error inesperado - {str(e)}")
                resultados['fallidos'] += 1
                wb.close()
                return resultados

        with transaction.atomic():
            for codigo_oc, datos_orden in ordenes_dict.items():
                # Verificar si la orden ya existe
                if OrdenCompra.objects.filter(codigo_oc=codigo_oc).exists():
                    resultados['errores'].append(f"La orden de compra '{codigo_oc}' ya existe en el sistema")
                    resultados['fallidos'] += 1
                    wb.close()
                    return resultados

                orden = OrdenCompra.objects.create(
                    codigo_oc=codigo_oc,
                    cliente=datos_orden['cliente'],
                    proyecto=datos_orden['proyecto'],
                    fecha_solicitud=datos_orden['fecha_solicitud']
                )

                for producto_data in datos_orden['productos']:
                    ProductoSolicitado.objects.create(
                        orden=orden,
                        producto=producto_data['producto'],
                        cantidad=producto_data['cantidad'],
                        descripcion=producto_data['descripcion']
                    )

                resultados['exitosos'] += 1

        wb.close()
        return resultados

    except Exception as e:
        resultados['errores'].append(f"Error al procesar el archivo: {str(e)}")
        resultados['fallidos'] += 1
        return resultados
import platform
import shutil
from django.shortcuts import render, get_object_or_404, redirect
from .models import Remision, DetalleRemision
from apps.utils.permisos import requiere_permiso
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.core.paginator import Paginator
from django.contrib import messages
from django.views.decorators.http import require_POST
import base64
import os
from django.conf import settings
import io
import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import subprocess
import tempfile
from openpyxl.worksheet.page import PageMargins


@requiere_permiso('ver_remisiones')
def lista_remisiones(request):
    buscar = request.GET.get('buscar', '')
    estado_facturacion = request.GET.get('estado_facturacion', '')
    
    remisiones = Remision.objects.all()
    
    if buscar:
        remisiones = remisiones.filter(numero_remision__icontains=buscar)
    
    if estado_facturacion:
        remisiones = remisiones.filter(estado_facturacion=estado_facturacion)
    
    remisiones = remisiones.order_by('-fecha_remision')
    
    paginator = Paginator(remisiones, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'remisiones/lista_remisiones.html', {
        'remisiones': page_obj,
        'page_obj': page_obj,
        'buscar': buscar,
        'estado_facturacion': estado_facturacion
    })


@requiere_permiso('ver_remisiones')
def detalle_remision(request, remision_id):
    remision = get_object_or_404(Remision, id=remision_id)
    detalles = remision.detalles.filter(despacho__reintegro=False)
    return render(request, 'remisiones/detalle_remision.html', {
        'remision': remision,
        'detalles': detalles
    })


@requiere_permiso('ver_remisiones')
@require_POST
def cambiar_estado_facturacion(request, remision_id):
    remision = get_object_or_404(Remision, id=remision_id)
    nuevo_estado = request.POST.get('estado')
    
    if nuevo_estado in ['pendiente', 'cerrado']:
        remision.estado_facturacion = nuevo_estado
        remision.save()
        messages.success(request, f'Estado de facturación actualizado a {nuevo_estado}')
    else:
        messages.error(request, 'Estado inválido')
    
    return redirect('lista_remisiones')


@requiere_permiso('ver_remisiones')
def subir_factura(request, remision_id):
    remision = get_object_or_404(Remision, id=remision_id)
    
    if request.method == 'POST':
        archivo = request.FILES.get('factura')
        if archivo:
            archivo_base64 = base64.b64encode(archivo.read()).decode('utf-8')
            remision.factura_base64 = archivo_base64
            remision.factura_nombre = archivo.name
            remision.estado_facturacion = 'cerrado'
            remision.save()
            messages.success(request, 'Factura subida exitosamente')
        else:
            messages.error(request, 'No se seleccionó ningún archivo')
        return redirect('lista_remisiones')
    
    return render(request, 'remisiones/subir_factura.html', {'remision': remision})


@requiere_permiso('ver_remisiones')
def ver_factura(request, remision_id):
    remision = get_object_or_404(Remision, id=remision_id)
    
    if not remision.factura_base64:
        messages.error(request, 'No hay factura disponible')
        return redirect('lista_remisiones')
    
    archivo_data = base64.b64decode(remision.factura_base64)
    
    extension = remision.factura_nombre.lower().split('.')[-1] if remision.factura_nombre else 'pdf'
    content_types = {
        'pdf': 'application/pdf',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png'
    }
    content_type = content_types.get(extension, 'application/octet-stream')
    
    response = HttpResponse(archivo_data, content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{remision.factura_nombre or "factura.pdf"}"'
    return response

@requiere_permiso('ver_remisiones')
def descargar_remision_excel(request, remision_id):
    remision = get_object_or_404(Remision, id=remision_id)
    detalles = remision.detalles.filter(despacho__reintegro=False)
    nombre_receptor = request.GET.get('nombre_receptor', '')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hoja1'
    
    # Datos básicos
    ws.cell(row=1, column=1, value=f"Remisión: {remision.numero_remision}")
    ws.cell(row=2, column=1, value=f"Fecha: {remision.fecha_remision.strftime('%d/%m/%Y')}")
    
    # Headers
    headers = ['Referencia', 'Descripción', 'Cantidad']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    
    # Datos
    row = 5
    for det in detalles:
        d = det.despacho
        prod_solic = getattr(d, 'producto_solicitado', None)
        if prod_solic:
            ws.cell(row=row, column=1, value=getattr(getattr(prod_solic, 'producto', None), 'referencia', ''))
            ws.cell(row=row, column=2, value=getattr(prod_solic, 'descripcion', ''))
            ws.cell(row=row, column=3, value=getattr(d, 'cantidad', ''))
            row += 1
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    response = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="remision_{remision.numero_remision}.xlsx"'
    return response


@requiere_permiso('ver_remisiones')
def descargar_remision_pdf(request, remision_id):
    remision = get_object_or_404(Remision, id=remision_id)
    detalles = remision.detalles.filter(despacho__reintegro=False)
    
    # Crear PDF simple
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="remision_{remision.numero_remision}.pdf"'
    
    # Contenido básico del PDF
    content = f"Remisión: {remision.numero_remision}\nFecha: {remision.fecha_remision.strftime('%d/%m/%Y')}\n\nDetalles:\n"
    for det in detalles:
        d = det.despacho
        prod_solic = getattr(d, 'producto_solicitado', None)
        if prod_solic:
            content += f"- {getattr(getattr(prod_solic, 'producto', None), 'referencia', '')}: {getattr(d, 'cantidad', '')}\n"
    
    response.write(content.encode('utf-8'))
    return response